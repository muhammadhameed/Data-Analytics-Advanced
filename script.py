import datetime
import pandas as pd
import json
import csv
import psycopg2
import requests
import math
import openpyxl
workbook = openpyxl.Workbook()
sheet = workbook.active
url = 'http://localhost:3001/candidate/create-candidate'
headers = {
    'accept': 'application/json',
    'Content-Type': 'application/json'
}
can = []
Pakistan =  [
        'Riphah International University,Islamabad',
        'Capital University of Science & Technology,Islamabad',
        'Quaid-i-Azam University,Islamabad',
        'Pakistan Institute of Engineering and Applied Sciences,Islamabad',
        'National Defence University, Pakistan,Islamabad',
        'Allama Iqbal Open University,Islamabad',
        'International Islamic University, Islamabad,Islamabad',
        'National University of Sciences and Technology, Pakistan,Islamabad',
        'COMSATS University,Islamabad',
        'National University of Computer and Emerging Sciences,Islamabad',
        'Bahria University,Islamabad',
        'Sir Syed CASE Institute of Technology,Islamabad',
        'Federal Urdu University of Arts, Science and Technology,Islamabad',
        'Institute of Space Technology,Islamabad',
        'Foundation University, Islamabad,Islamabad',
        'National University of Modern Languages,Islamabad',
        'Air University, Islamabad,Islamabad',
        'Shifa Tameer-e-Millat University,Islamabad',
        'Shaheed Zulfiqar Ali Bhutto Medical University,Islamabad',
        'Muslim Youth University,Islamabad',
        'National University of Technology,Islamabad',
        'Pakistan Institute of Development Economics,Islamabad',
        'National Skills University,Islamabad',
        'Ibadat International University,Islamabad',
        'University of Balochistan,Quetta',
        'Balochistan University of Engineering and Technology,Khuzdar',
        'Al-Hamd Islamic University,Quetta',
        'Balochistan University of Information Technology, Engineering and Management Sciences,Quetta',
        "Sardar Bahadur Khan Women's University, Quetta",
        'Lasbela University of Agriculture, Water and Marine Sciences,Lasbela',
        'University of Turbat,Turbat ',
        'University of Loralai,Loralai',
        'Mir Chakar Khan Rind University,Sibi',
        'University of Gwadar,Gwadar',
        'Islamia College University,Peshawar',
        'Pakistan Military Academy,Abbotabad',
        'Pakistan Air Force Academy,Risalpur',
        'University of Peshawar,Peshawar',
        'Gomal University,Dera Ismail Khan',
        'University of Engineering and Technology, Peshawar,Peshawar',
        'University of Agriculture, Peshawar,Peshawar',
        'Preston University,Kohat',
        'CECOS University of Information Technology and Emerging Sciences,Peshawar',
        'Ghulam Ishaq Khan Institute of Engineering Sciences and Technology,Swabi',
        'Kohat University of Science and Technology,Kohat',
        'University of Malakand,Chakdara',
        'Qurtuba University,Peshawar',
        'Sarhad University of Science and Information Technology,Peshawar',
        'City University of Science and Information Technology, Peshawar,Peshawar',
        'Hazara University,Mansehra',
        'Gandhara University,Peshawar',
        'Northern University, Nowshera,Nowshera',
        'Institute of Management Sciences Peshawar,Peshawar',
        'University of Science and Technology Bannu,Bannu',
        'Shaheed Benazir Bhutto Women University,Peshawar',
        'Khyber Medical University,Peshawar',
        'Abasyn University,Peshawar',
        'Abdul Wali Khan University Mardan,Mardan',
        'Shaheed Benazir Bhutto University, Sheringal,Upper Dir',
        'University of Swat,Swat',
        'Bacha Khan University,Charsadda',
        'University of Haripur,Haripur',
        'IQRA National University,Peshawar',
        'Khushal Khan Khattak University,Karak',
        'University of Swabi,Swabi',
        'University of Buner,Buner',
        'Abbottabad University of Science and Technology,Abbottabad',
        'Women University Mardan,Mardan',
        'Women University Swabi,Swabi',
        'University of Technology, Nowshera,Nowshera',
        'FATA University,Akhorwal',
        'University of Chitral,Chitral',
        'University of Engineering and Technology, Mardan,Mardan',
        'University of Agriculture, Dera Ismail Khan,Dera Ismail Khan',
        'University of Lakki Marwat,Lakki Marwat',
        'Pak-Austria Fachhochschule: Institute of Applied Sciences and Technology,Haripur',
        'University of Dir,Timergara',
        'King Edward Medical University,Lahore',
        'Faisalabad Medical University,Faisalabad',
        'Forman Christian College,Lahore',
        'University of Veterinary and Animal Sciences,Lahore',
        'University of the Punjab,Lahore',
        'Punjab Tianjin University of Technology,Lahore',
        'Kinnaird College for Women University,Lahore',
        'University of Engineering and Technology, Lahore,Lahore',
        'Lahore College for Women University,Lahore',
        'Government College University, Faisalabad,Faisalabad',
        'National College of Arts,Lahore',
        'University of Agriculture, Faisalabad,Faisalabad',
        'Namal Institute,Mianwali',
        'Fatima Jinnah Medical University,Lahore',
        'National Textile University,Faisalabad',
        'Pir Mehr Ali Shah Arid Agriculture University,Rawalpindi',
        'Bahauddin Zakariya University,Multan',
        'The Islamia University of Bahawalpur,Bahawalpur',
        'University of Engineering and Technology, Taxila,Taxila',
        'Lahore University of Management Sciences,Lahore',
        'NFC Institute of Engineering and Technology,Multan',
        'Institute of Management Sciences, Lahore,Lahore',
        'University of Management and Technology, Lahore,Lahore',
        'National College of Business Administration and Economics,Lahore',
        'Government College University, Lahore,Lahore',
        'Fatima Jinnah Women University,Rawalpindi',
        'University of Sargodha,Sargodha',
        'University of Health Sciences, Lahore,Lahore',
        'University of Education,Lahore',
        'GIFT University,Gujranwala',
        'Hajvery University,Lahore',
        'University of Central Punjab,Lahore',
        'University of Faisalabad,Faisalabad',
        'University of Lahore,Lahore',
        'Beaconhouse National University,Lahore',
        'University of South Asia,Lahore',
        'University of Gujrat,Gujrat',
        'Superior University,Lahore',
        'Minhaj University, Lahore,Lahore',
        'HITEC University,Taxila',
        'University of Wah,Wah',
        'Pakistan Institute of Fashion and Design,Lahore',
        'Women University Multan,Multan',
        'Institute of Southern Punjab,Multan',
        'Qarshi University,Lahore',
        'Government College Women University, Sialkot,Sialkot',
        'Government Sadiq College Women University,Bahawalpur',
        'Ghazi University,Dera Ghazi Khan',
        'Government College Women University, Faisalabad,Faisalabad',
        'Information Technology University (Lahore),Lahore',
        'Muhammad Nawaz Sharif University of Agriculture,Multan',
        'Muhammad Nawaz Sharif University of Engineering and Technology,Multan',
        'Virtual University of Pakistan,Lahore',
        'Lahore Garrison University,Lahore',
        'Cholistan University of Veterinary and Animal Sciences,Bahawalpur',
        'Khawaja Fareed University of Engineering and Information Technology,Rahim Yar Khan',
        'University of Engineering and Technology, Rasul,Mandi Bahauddin',
        'University of Sahiwal,Sahiwal',
        'Institute for Art and Culture,Lahore',
        'University of Mianwali,Mianwali',
        'Thal University,Bhakkar',
        'Green International University,Lahore',
        'Lahore Institute of Science and Technology,Lahore',
        'Grand Asian University Sialkot,Sialkot',
        'Liaquat University of Medical and Health Sciences,Jamshoro',
        'Sindh Madressatul Islam University,Karachi',
        'NED University of Engineering and Technology,Karachi',
        'Dow University of Health Sciences,Karachi',
        'University of Sindh,Jamshoro',
        'University of Karachi,Karachi',
        'Institute of Business Administration, Karachi',
        'Dawood University of Engineering and Technology,Karachi',
        'Mehran University of Engineering and Technology,Jamshoro',
        'University of Okara,Okara',
        'University of Jhang,Jhang',
        'NUR International University,Lahore',
        'University of Sialkot,Sialkot',
        'Lahore School of Economics,Lahore',
        'Rawalpindi Medical University,Rawalpindi',
        'Nishtar Medical University,Multan',
        'National University of Medical Sciences,Rawalpindi',
        'University of Home Economics Lahore,Lahore',
        'Mir Chakar Khan Rind University of Technology,Dera Ghazi Khan',
        'Rawalpindi Women University,Rawalpindi',
        'University of Narowal,Narowal',
        'Al-Qadir University,Sohawa',
        'Baba Guru Nanak University,Nankana Sahib',
        'University of Chakwal,Chakwal',
        'Kohsar University Murree,Murree',
        'Government Viqar-un-Nisa Women University,Rawalpindi',
        'Pakistan Naval Academy,Karachi',
        'Shah Abdul Latif University,Khairpur',
        'Quaid-e-Awam University of Engineering, Science and Technology,Benazirabad',
        'Sindh Agriculture University,Tandojam',
        'Indus Valley School of Art and Architecture,Karachi',
        'Baqai Medical University,Karachi',
        'Hamdard University,Karachi',
        'Commecs institute of business and emerging sciences,Karachi',
        'Sir Syed University of Engineering and Technology,Karachi',
        'Sukkur Institute of Business Administration (IBA) University,Sukkur',
        'Textile Institute of Pakistan,Karachi',
        'Institute of Business Management,Karachi',
        'Shaheed Zulfiqar Ali Bhutto Institute of Science and Technology,Karachi',
        'Preston Institute of Management Sciences and Technology,Karachi',
        'Sindh Institute of Medical Sciences,Karachi',
        'Pir Abdul Qadir Shah Jeelani Institute of Medical Sciences,Khairpur',
        'Qalandar Shahbaz University of Modern Sciences,Karachi',
        'Benazir Bhutto Shaheed University of Technology and Skill Development,Khairpur',
        'Mirpur University of Science and Technology,Mirpur',
        'University of Azad Jammu and Kashmir,Muzaffarabad',
        'University of Poonch,Rawalakot',
        'Al-Khair University,Mirpur',
        'Mohi-ud-Din Islamic University,Nerian Sharif',
        'Women University of Azad Jammu and Kashmir Bagh,Bagh',
        'University of Kotli,Kotli',
        'Isra University,Hyderabad',
        'Karachi Institute of Economics and Technology,Karachi',
        'Greenwich University, Karachi,Karachi',
        'Jinnah University for Women,Karachi',
        'Iqra University,Karachi',
        'Dadabhoy Institute of Higher Education,Karachi',
        'Ilma University,Karachi',
        'Indus University,Karachi',
        'University of EAST,Hyderabad',
        'Aga Khan University,Karachi',
        'Shaheed Mohtarma Benazir Bhutto Medical University,Larkana',
        'Muhammad Ali Jinnah University,Karachi',
        'Karachi School for Business and Leadership,Karachi',
        'Habib University,Karachi',
        'Benazir Bhutto Shaheed University,Karachi',
        'Shaheed Benazir Bhutto University, Benazirabad,Benazirabad',
        'KASB Institute of Technology,Karachi',
        'Jinnah Sindh Medical University,Karachi',
        'Shaheed Zulfiqar Ali Bhutto University of Law,Karachi',
        'DHA Suffa University,Karachi',
        'Nazeer Hussain University,Karachi',
        'Peoples University of Medical and Health Sciences for Women,Benazirabad',
        'Shaheed Benazir Bhutto University of Veterinary and Animal Sciences,Benazirabad',
        'Newports Institute of Communications and Economics,Karachi',
        'Shaheed Benazir Bhutto City University,Karachi',
        'Shaheed Benazir Bhutto Dewan University,Karachi',
        'Hyderabad Institute of Arts, Science and Technology,Hyderabad',
        'Begum Nusrat Bhutto Women University,Sukkur',
        'University of Sufism and Modern Sciences,Bhit Shah',
        'Government College University Hyderabad,Hyderabad',
        'Shaikh Ayaz University,Shikarpur',
        'Ziauddin University,Karachi',
        'Salim Habib University,Karachi',
        'Karakoram International University,Gilgit',
        'Baltistan University,Skardu',
        'LGS International Degree Programme, Lahore',
        'LGS International Degree Programme, Islamabad',
        'LGS International Degree Programme, Karachi',
        'LGS International Degree Programme, Peshawar',
        'LGS International Degree Programme, Gujrat',
        'Universal College of Lahore, Lahore',
        'Other',
        'Abasyn University, Peshawar',
        'Abbottabad University of Science and Technology, Abbottabad',
        'Abdul Wali Khan University, Mardan',
        'Aga Khan University, Karachi',
        'Air University, Islamabad',
        'Air University (Pakistan Air Force), Islamabad',
        'Al-Hamd Islamic University, Sariab',
        'Al-Khair University, Rawalpindi',
        'Allama Iqbal Open University , Islamabad',
        'Bacha Khan University, Khyber Pakhtunkhwa',
        'Bahauddin Zakariya University, Multan',
        'Bahria University, Islamabad',
        'Balochistan University of Engineering and Technology, Khuzdar',
        'Baqai Medical University, Karachi',
        'Beaconhouse National University, Lahore',
        'CECOS University of IT and Emerging Sciences, Peshawar',
        'City University of Science & Information Technology, Peshawar',
        'City University of Science and IT, Peshawar',
        'COMSATS Institute of Information Technology Islamabad, Islamabad',
        'DHA Suffa University, , Karachi',
        'Dow University of Health Sciences, Saddar Town',
        'Fast University, Peshawar Campus, Islamabad,',
        'Fatima Jinnah Women University, Rawalpindi',
        'Federal Urdu University of Arts, Sciences and Technology, Karachi',
        'Federal Urdu University of Arts, Sciences and Technology Islamabad, Islamabad',
        'Foundation University, Islamabad',
        'Gandhara University, Peshawar',
        'Ghazi University, Punjab,',
        'GIFT University, Gujranwala',
        'Global Institute, Lahore, Lahore',
        'Gomal University, Dera Ismail Khan',
        'Government College University Faisalabad, Faisalabad',
        'Government College University Lahore, Lahore',
        'Greenwich University, Karachi',
        'Habib University, Karachi',
        'Hajvery University, Lahore',
        'Hamdard University, Karachi',
        'Hazara University, Dhodial',
        'HITEC University, Taxila',
        'Indus University, Karachi',
        'Information Technology University, Lahore, Lahore',
        'International Islamic University, Islamabad, Islamabad',
        'IQRA National University, Peshawar,',
        'Iqra University, Karachi',
        'Islamia College University, Peshawar',
        'Islamia University, Bahawalpur',
        'Isra University, Hyderabad',
        'Jinnah University for Women, Karachi',
        'Karakurum International University, Gilgit',
        'KASB Institute of Technology, Karachi',
        'Khushal Khan Khattak University, Khyber-Pakhtunkhwa',
        'Khyber Medical University, , Peshawar',
        'King Edward Medical University, Lahore',
        'Kohat University of Science & Technology, Kohat',
        'Lahore College for Women University, Lahore',
        'Lahore Garrison University, Lahore,',
        'Lahore University of Management Sciences, Lahore',
        'Lasbela University of Agriculture, Water and Marine Sciences, Lasbela',
        'Liaquat University of Medical & Health Sciences, Jamshoro',
        'Mehran University of Engineering & Technology, Jamshoro',
        'Minhaj University Lahore, Lahore',
        'Mirpur University of Science and Technology, Mirpur,',
        'Mohammad Ali Jinnah University, Karachi',
        'Mohi-ud-Din Islamic University, Nerian Sharif',
        'Muhammad Nawaz Sharif University of Agriculture, Multan',
        'Muslim Youth University, Islamabad',
        'National Defence University, Islamabad',
        'National Textile University, Faisalabad',
        'National University of Computer & Emerging Sciences (FAST), Islamabad',
        'National University of Modern Languages, Islamabad',
        'National University of Sciences & Technology, Islamabad',
        'Nazeer Hussain University, Karachi',
        'NED University of Engineering and Technology, Karachi',
        'Northern University, Nowshera',
        'Nur International University, Lahore',
        'NWFP Agriculture University, Peshawar',
        'NWFP University of Engineering & Technology, Peshawar',
        'Pakistan Institute of Engineering and Applied Sciences, Islamabad',
        'Pir Mehr Ali Shah Arid Agriculture University, Punjab,',
        'Preston University, Karachi',
        'Qalandar Shahbaz University of Modern Sciences, Karachi',
        'Qarshi University, Punjab,',
        'Quaid-e-Awam University of Engineering, Science & Technology, Nawabshah',
        'Quaid-i-Azam University, Islamabad',
        'Qurtaba University of Science & Information Technology, Dera Ismail Khan',
        'Qurtuba University, Khyber Pakhtunkhwa,',
        'Riphah International University, Islamabad',
        "Sardar Bahadur Khan Women's University, Quetta",
        'Sarhad University of Science & Information Technology, Peshawar',
        'Sarhad University of Science and IT, Peshawar,',
        'Shah Abdul Latif University, Khairpur',
        'Shaheed Benazir Bhutto City University, Karachi,',
        'Shaheed Benazir Bhutto University, Sheringal, Sheringal',
        'Shaheed Benazir Bhutto Women University, Peshawar',
        'Shaheed Zulfiqar Ali Bhutto Medical University, Islamabad',
        'Shifa Tameer-e-Millat University, Islamabad',
        'Sindh Agriculture University, Tandojam',
        'Sir Syed University of Engineering & Technology, Karachi',
        'Superior University, Garden Town',
        'The Balochistan University of Information Technology, Engineering, and Management Sciences, Quetta',
        'University of Agriculture, Faisalabad, Faisalabad',
        'University of Agriculture, Peshawar, Peshawar',
        'University of Arid Agriculture, Rawalpindi',
        'University of Azad Jammu and Kashmir, Muzaffarabad',
        'University of Balochistan, Quetta',
        'University of Central Punjab, Lahore',
        'University of East, Hyderabad',
        'University of Education, Lahore',
        'University of Engineering & Technology, Lahore, Lahore',
        'University of Engineering & Technology, Taxila, Taxila',
        'University of Engineering and Technology, Peshawar, Peshawar',
        'University of Faisalabad, Faisalabad',
        'University of FATA, FR Kohat',
        'University of Gujrat, Gujrat',
        'University of Haripur, Khyber Pakhtunkhwa,',
        'University of Health Sciences, Lahore, Lahore',
        'University of Information Technnology, Engineering & Management Sciences, Quetta',
        'University of Jhang, Jhang',
        'University of Karachi, Karachi',
        'University of Lahore, Lahore',
        'University of Loralai, Balochistan',
        'University of Malakand, Chakdara',
        'University of Management and Technology, Lahore',
        'University of Munawwar-ul-islam, Gujrat',
        'University of Okara, Okara',
        'University of Peshawar, Pakistan, Peshawar',
        'University of Poonch, Rawalakot',
        'University of Sahiwal, Sahiwal',
        'University of Sargodha, Sargodha',
        'University of Science & Technology Bannu, Bannu',
        'University of Science and Technology, Pakhtunkhwa',
        'University of Sindh, Jamshoroo',
        'University of South Asia, Lahore',
        'University of Swabi, Swabi',
        'University of Swat, Odigram,',
        'University of the Punjab, Lahore',
        'University of Turbat, Turbat',
        'University of Veterinary and Animal Sciences, Lahore',
        'University of Wah, Wah Cantt.',
        'Virtual University of Pakistan, Lahore',
        'Women University Mardan, Mardan',
        'Women University Multan, Multan,',
        'Women University of Azad Jammu and Kashmir, Bagh, Kotli',
        'Ziauddin University, Karachi',
      ]
rejected = []
# Load data from Excel sheet
def replace_nan_values(candidate):
    cleaned_candidate = {}
    for key, value in candidate.items():
        if isinstance(value, dict):
            cleaned_value = replace_nan_values(value)  # Recursively replace nested dictionaries
            cleaned_candidate[key] = cleaned_value
        elif isinstance(value, list):
            cleaned_list = []
            for item in value:
                if isinstance(item, dict):
                    cleaned_item = replace_nan_values(item)  # Recursively replace nested dictionaries in lists
                    cleaned_list.append(cleaned_item)
                elif isinstance(item, float) and math.isnan(item):
                    cleaned_list.append('')  # Replace NaN values in lists with ''
                else:
                    cleaned_list.append(item)
            cleaned_candidate[key] = cleaned_list
        elif isinstance(value, float) and math.isnan(value):
            cleaned_candidate[key] = ''  # Replace NaN values with ''
        else:
            cleaned_candidate[key] = value
    return cleaned_candidate
f = open('output.txt','w+')
df = pd.read_excel('fin2.xlsx')

print(df['jobFromDate'][0])
j = 1
# Iterate over each row in the dataframe
for i, row in df.iterrows():
    # Create a dictionary for the candidate's details
    candidate = {
        "firstName": row['firstName'],
        "lastName": row['lastName'],
        "email": row['email'],
        "nationality": row['nationality'],
        "country": row['country'],
        "city": row['city'],
        "gender": row['gender'],
        "experienceYears": row['experienceYears'],
        "resumeUrl": row['resumeUrl'],
        "linkedInUrl": row['linkedInUrl'],
        "positionToApply": 'Analyst',
        "availableStartDate": '1111-11-11T11:11:11',
        "universities": [
            {
                "location": 'Pakistan' if row['universityName'] in Pakistan else 'International',
                "universityName": row['universityName'],
                "qualification": row['qualification'],
                "areaOfStudy": row['areaOfStudy'],
                "gpa": row['gpa'],
                "fromDate": str(row['fromDate'])[:10]+'T'+str(row['fromDate'])[11:],
                "toDate": str(row['toDate'])[:10]+'T'+str(row['toDate'])[11:]
            }
        ],
        "schools": [
            {
                "location": 'Pakistan' if row['universityName'] in Pakistan else 'International',
                "schoolName": row['schoolName'],
                "qualification": row['schoolQualification'],
                "result": row['result'],
                "fromDate": str(row['fromDate'])[:10]+'T'+str(row['toDate'])[11:],
                "toDate": str(row['SchooltoDate'])[:10]+'T'+str(row['SchooltoDate'])[11:]
            }
        ],
        "workExperiences": [
            {
                "jobTitle": row['jobTitle'],
                "company": row['company'],
                "industry": 'xyz',
                "department": 'xyz',
                "currentlyWorking": row['currentlyWorking'],
                "fromDate": str(row['jobFromDate'])[:10]+'T'+str(row['jobFromDate'])[11:],
                "toDate": str(row['jobToDate'])[:10]+'T'+str(row['jobToDate'])[11:]
            }
        ],
        "consultingExperience": [],
        "consultingExperienceDescription": "",
        "candidatePhotograph": "dummy.com",
        "skills": row['skills'].split(","),
        "languages": row['languages'].split(","),
        "hearAboutIntellia": "",
        "knowAnyoneAtIntellia": row['knowAnyoneAtIntellia'],
        "referredBy": row['referredBy'],
        "workedAtIntellia": row['workedAtIntellia'],
        "previouslyWorkedEndDate": str(row['previouslyWorkedEndDate'])[:10]+'T'+str(row['previouslyWorkedEndDate'])[11:],
        "previouslyWorkedStartDate": str(row['previouslyWorkedStartDate'])[:10]+'T'+str(row['previouslyWorkedStartDate'])[11:],
        "zohoId": row['Candidate Id']
    }
    
    cleaned_candidate = replace_nan_values(candidate)
    # Convert dictionary to JSON string

    candidate_json = json.dumps(cleaned_candidate, default=str)
    print(candidate_json)
    response = requests.post(url, headers=headers, data=candidate_json)
    print(response.status_code)
    print(response.text)
    
    if response.status_code != 201:
        if response.status_code != 422:
            rejected.append((candidate_json,response.text))
            sheet.cell(row=j, column=1, value=candidate_json)
            sheet.cell(row=j, column=2, value=response.text)
            j+=1
               
    print('HERE')
    f.write(candidate_json)
    f.write(',')
    can.append(candidate_json)
    
workbook.save("problems.xlsx")
print(len(rejected))
# Convert the list of JSON objects to a comma-separated string


# Write the comma-separated string to the TXT file
